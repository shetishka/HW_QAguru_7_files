from paths import *
import csv
from pypdf import PdfReader
from zipfile import ZipFile
from openpyxl import load_workbook


# – Запаковать кодом в zip архив несколько разных файлов: pdf, xlsx, csv;
# – Положить его в ресурсы;
# – Реализовать чтение и проверку содержимого каждого файла из архива не распаковывая сам архив

def test_archive_csv():
    with ZipFile(ARCHIVE) as zip_file:
        with zip_file.open('CSV_file.csv') as csv_file:
            content = csv_file.read().decode('utf-8-sig')
            csvreader = list(csv.reader(content.splitlines()))
            second_row = csvreader[1]

            assert second_row[1] == '5'


def test_archive_xlsx():
    with ZipFile(ARCHIVE) as zip_file:
        with zip_file.open('file_example_XLSX_50.xlsx') as xlsx_file:
            workbook = load_workbook(xlsx_file)
            sheet = workbook.active

            assert sheet.cell(row=1, column=2).value == "First Name"


def test_archive_pdf():
    with ZipFile(ARCHIVE) as zip_file:
        with zip_file.open('Polozhenie_Pesni_Dlya_Vsekh_2024.pdf') as pdf_file:
            reader = PdfReader(pdf_file)
            page = reader.pages[0]
            text = page.extract_text()

            assert "12-13 октября  2024 года " in text
